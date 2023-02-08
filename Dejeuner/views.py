from django.shortcuts import render
from django.http import HttpResponseRedirect
from .models import PlaceDispo, ContactModel, CompteurMenu
from openpyxl import load_workbook
import os

def dejeuner(request):

    horaires=PlaceDispo.objects.all()

    context = {
        'horaires': horaires,
    }

    if request.method == 'POST':
            nom= request.POST.get('Nom').lower()
            prenom= request.POST.get('Prenom').lower()
            adresse= request.POST.get('Adresse')
            message= request.POST.get('Message')
            vieux=request.POST.get('Vieux')
            couvert=request.POST.get('Couvert')
            menu=request.POST.get('Menu')
            num = request.POST.get('Num')
            boisson= request.POST.get('Boisson')
            horaire = request.POST.get('Horaire')




            Users = ContactModel.objects.all()
            Horaire_list = PlaceDispo.objects.all()

            list_prenom, list_nom = ['PiVoT'], ['PiVoT']
            #fonction qui enleve les majuscules dans les strings

            for i in Users:
                list_prenom.append(i.Prenom)
                list_nom.append(i.Nom)
            for i in range(0,len(list_nom)):
                if nom.lower() == list_nom[i]:
                    for j in range(0,len(list_prenom)):
                        if list_prenom[j]==prenom.lower():
                            return HttpResponseRedirect('/erreur1/')
                            break





            wb_cotisant = load_workbook('Dejeuner/Cotisants-BDE-respo-num.xlsx')
            ws_cotisant = wb_cotisant.active
            cotisant = False
            for i in range(2, 1507):
                if ws_cotisant[f'A{i}'].value.lower() == nom.lower() and  ws_cotisant[f'B{i}'].value.lower() == prenom.lower():
                    cotisant = True
                    break

            if cotisant == False:
                wb_cotisant.save('Dejeuner/Cotisants-BDE-respo-num.xlsx')
                return HttpResponseRedirect('/erreur1/')


            wb_cotisant.save('Dejeuner/Cotisants-BDE-respo-num.xlsx')


            for i in Horaire_list:
                if i.Horaire==horaire:
                    i.Count-=1
                    i.save()
                    break

            compteur_menus = CompteurMenu.objects.all()

            for menu_display in compteur_menus:
                if menu_display.menu == menu:
                    menu_display.compteur += 1
                    menu_display.save()

            if vieux=='on':
                vieux=True
            elif vieux==None:
                vieux=False
            if couvert=='on':
                couvert=True
            elif couvert==None:
                couvert=False


            #dernier test --> matcher liste prénoms + nom école : fait ok

            wb = load_workbook('Dejeuner/petitdej.xlsx')
            ws = wb[f'{horaire}']
            ws[f'A{ws["L1"].value}'].value = adresse
            ws[f'B{ws["L1"].value}'].value = 'Evry'
            ws[f'C{ws["L1"].value}'].value = 91000
            ws[f'D{ws["L1"].value}'].value = prenom
            ws[f'E{ws["L1"].value}'].value = nom
            ws[f'F{ws["L1"].value}'].value = num
            ws[f'G{ws["L1"].value}'].value = vieux
            ws[f'H{ws["L1"].value}'].value = couvert
            ws[f'I{ws["L1"].value}'].value = menu
            ws[f'J{ws["L1"].value}'].value = boisson
            ws[f'K{ws["L1"].value}'].value = message

            ws['L1'].value += 1

            wb.save('Dejeuner/petitdej.xlsx')


            NewComamande= ContactModel.objects.create(Nom= nom, Prenom=prenom, Adresse=adresse, Message=message, Vieux=vieux,Couvert=couvert,Horaire=horaire, Menu=menu, Num= num, Boisson=boisson)
            NewComamande.save()

            #form.save()
            return HttpResponseRedirect('/redirectionDej/')


    return render(request, "dejeuner/dejeuner.html", context)

def redirection(request):
    return render(request, "dejeuner/redirection1.html")

def erreur1(request):
    return render(request, "dejeuner/erreur1.html")

def menu(request):
    return render(request, "dejeuner/menu.html")

